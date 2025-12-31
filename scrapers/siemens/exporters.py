"""
Export functions for Siemens vulnerability data.
"""
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import SiemensVulnerability


def export_to_json(vulnerabilities: list[SiemensVulnerability], output_path: str) -> None:
    """Export vulnerabilities to JSON file."""
    data = {
        "metadata": {
            "source": "Siemens ProductCERT",
            "scraped_at": datetime.utcnow().isoformat() + "Z",
            "total_count": len(vulnerabilities),
            "url": "https://cert-portal.siemens.com/productcert/html/"
        },
        "vulnerabilities": [v.to_dict() for v in vulnerabilities]
    }
    
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"✓ Exported {len(vulnerabilities)} vulnerabilities to {output_path}")


def export_to_excel(vulnerabilities: list[SiemensVulnerability], output_path: str) -> None:
    """Export vulnerabilities to Excel file with multiple sheets."""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Severity colors
    severity_fills = {
        "Critical": PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid"),
        "High": PatternFill(start_color="ea580c", end_color="ea580c", fill_type="solid"),
        "Medium": PatternFill(start_color="ca8a04", end_color="ca8a04", fill_type="solid"),
        "Low": PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid"),
        "Unknown": PatternFill(start_color="6b7280", end_color="6b7280", fill_type="solid"),
    }
    
    # Sheet 1: Overview
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    overview_headers = [
        "SSA ID", "Title", "CVSS v3.1", "CVSS v4.0", "Severity",
        "CVE IDs", "Publication Date", "Last Update", "Version", "Summary"
    ]
    
    for col, header in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, vuln in enumerate(vulnerabilities, 2):
        ws_overview.cell(row=row, column=1, value=vuln.ssa_id).border = thin_border
        ws_overview.cell(row=row, column=2, value=vuln.title).border = thin_border
        ws_overview.cell(row=row, column=3, value=vuln.cvss_v3_score).border = thin_border
        ws_overview.cell(row=row, column=4, value=vuln.cvss_v4_score).border = thin_border
        
        severity_cell = ws_overview.cell(row=row, column=5, value=vuln.severity)
        severity_cell.border = thin_border
        severity_cell.fill = severity_fills.get(vuln.severity, severity_fills["Unknown"])
        severity_cell.font = Font(color="FFFFFF", bold=True)
        
        ws_overview.cell(row=row, column=6, value=", ".join(vuln.cve_ids)).border = thin_border
        ws_overview.cell(row=row, column=7, value=vuln.publication_date).border = thin_border
        ws_overview.cell(row=row, column=8, value=vuln.last_update).border = thin_border
        ws_overview.cell(row=row, column=9, value=vuln.version).border = thin_border
        ws_overview.cell(row=row, column=10, value=vuln.summary[:500] + "..." if len(vuln.summary) > 500 else vuln.summary).border = thin_border
    
    # Adjust column widths for overview
    overview_widths = [15, 60, 12, 12, 12, 30, 15, 15, 10, 80]
    for col, width in enumerate(overview_widths, 1):
        ws_overview.column_dimensions[get_column_letter(col)].width = width
    
    # Sheet 2: Affected Products
    ws_products = wb.create_sheet("Affected Products")
    
    product_headers = ["SSA ID", "Product", "Affected Versions", "Remediation", "CVE ID"]
    for col, header in enumerate(product_headers, 1):
        cell = ws_products.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    product_row = 2
    for vuln in vulnerabilities:
        for product in vuln.affected_products:
            if isinstance(product, dict):
                ws_products.cell(row=product_row, column=1, value=vuln.ssa_id).border = thin_border
                ws_products.cell(row=product_row, column=2, value=product.get('product', '')).border = thin_border
                ws_products.cell(row=product_row, column=3, value=product.get('versions', '')).border = thin_border
                ws_products.cell(row=product_row, column=4, value=product.get('remediation', '')).border = thin_border
                ws_products.cell(row=product_row, column=5, value=product.get('cve_id', '')).border = thin_border
            else:
                ws_products.cell(row=product_row, column=1, value=vuln.ssa_id).border = thin_border
                ws_products.cell(row=product_row, column=2, value=product.product).border = thin_border
                ws_products.cell(row=product_row, column=3, value=product.versions).border = thin_border
                ws_products.cell(row=product_row, column=4, value=product.remediation).border = thin_border
                ws_products.cell(row=product_row, column=5, value=product.cve_id or '').border = thin_border
            product_row += 1
    
    product_widths = [15, 40, 30, 60, 20]
    for col, width in enumerate(product_widths, 1):
        ws_products.column_dimensions[get_column_letter(col)].width = width
    
    # Sheet 3: CVE Mapping
    ws_cve = wb.create_sheet("CVE Mapping")
    
    cve_headers = ["SSA ID", "CVE ID", "Title", "CVSS v3.1", "Severity"]
    for col, header in enumerate(cve_headers, 1):
        cell = ws_cve.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    cve_row = 2
    for vuln in vulnerabilities:
        for cve_id in vuln.cve_ids:
            ws_cve.cell(row=cve_row, column=1, value=vuln.ssa_id).border = thin_border
            ws_cve.cell(row=cve_row, column=2, value=cve_id).border = thin_border
            ws_cve.cell(row=cve_row, column=3, value=vuln.title).border = thin_border
            ws_cve.cell(row=cve_row, column=4, value=vuln.cvss_v3_score).border = thin_border
            
            severity_cell = ws_cve.cell(row=cve_row, column=5, value=vuln.severity)
            severity_cell.border = thin_border
            severity_cell.fill = severity_fills.get(vuln.severity, severity_fills["Unknown"])
            severity_cell.font = Font(color="FFFFFF", bold=True)
            cve_row += 1
    
    cve_widths = [15, 20, 60, 12, 12]
    for col, width in enumerate(cve_widths, 1):
        ws_cve.column_dimensions[get_column_letter(col)].width = width
    
    # Sheet 4: Links
    ws_links = wb.create_sheet("Links")
    
    link_headers = ["SSA ID", "HTML URL", "CSAF URL", "PDF URL", "TXT URL"]
    for col, header in enumerate(link_headers, 1):
        cell = ws_links.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, vuln in enumerate(vulnerabilities, 2):
        ws_links.cell(row=row, column=1, value=vuln.ssa_id).border = thin_border
        ws_links.cell(row=row, column=2, value=vuln.html_url).border = thin_border
        ws_links.cell(row=row, column=3, value=vuln.csaf_url or '').border = thin_border
        ws_links.cell(row=row, column=4, value=vuln.pdf_url or '').border = thin_border
        ws_links.cell(row=row, column=5, value=vuln.txt_url or '').border = thin_border
    
    link_widths = [15, 70, 70, 70, 70]
    for col, width in enumerate(link_widths, 1):
        ws_links.column_dimensions[get_column_letter(col)].width = width
    
    # Save workbook
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    print(f"✓ Exported {len(vulnerabilities)} vulnerabilities to {output_path}")

